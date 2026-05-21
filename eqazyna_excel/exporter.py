from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

EXPORT_COLUMNS = [
    ("exported_at", "Время выгрузки"),
    ("created_at", "Дата создания"),
    ("document_number", "Номер документа"),
    ("applicant_bin", "БИН"),
    ("applicant_name", "Заявитель из e-Qazyna"),
    ("document_type", "Тип документа"),
    ("status", "Статус заявки"),
    ("egov_found", "eGov найдено"),
    ("egov_name_ru", "Наименование eGov RU"),
    ("egov_name_kz", "Наименование eGov KZ"),
    ("legal_address_ru", "Юридический адрес RU"),
    ("legal_address_kz", "Юридический адрес KZ"),
    ("region", "Регион по адресу"),
    ("region_confidence", "Уверенность региона"),
    ("leader_fio", "Руководитель"),
    ("oked_code", "ОКЭД код"),
    ("oked_name_ru", "Вид деятельности"),
    ("registration_date", "Дата регистрации"),
    ("company_status", "Статус юрлица"),
    ("egov_error", "Ошибка/примечание eGov"),
    ("search_2gis", "Поиск 2GIS"),
    ("search_google", "Поиск Google"),
    ("search_yandex", "Поиск Яндекс"),
    ("source_page", "Страница e-Qazyna"),
    ("source_url", "Источник e-Qazyna"),
    ("unique_key", "Технический ключ"),
]

README_ROWS = [
    ["Что это", "Выгрузка заявок e-Qazyna с обогащением по БИН через data.egov.kz."],
    ["Фильтр", "Тип документа: Заявка на разведку ТПИ; статус: Отправлено на рассмотрение / Принято."],
    ["Адрес", "Сначала берётся юридический адрес из data.egov.kz. Если eGov не нашёл БИН, адрес остаётся пустым."],
    ["Телефон", "Телефон автоматически не парсится. В таблице есть ссылки на 2GIS / Google / Яндекс для ручной проверки."],
    ["Регион", "Определяется по найденному юридическому адресу и словарю config/regions.yaml."],
]


def export_to_excel(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    info = wb.create_sheet("Описание")

    _write_info(info)
    _write_data(ws, rows)
    wb.save(output_path)
    return output_path


def _write_info(ws) -> None:
    ws.append(["Раздел", "Описание"])
    for row in README_ROWS:
        ws.append(row)
    ws.append(["Источники", "e-Qazyna: https://minerals.e-qazyna.kz/ru/guest/reestr/doc/list"])
    ws.append(["Источники", "data.egov.kz API: https://data.egov.kz/pages/samples"])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"


def _write_data(ws, rows: list[dict[str, Any]]) -> None:
    headers = [title for _, title in EXPORT_COLUMNS]
    ws.append(headers)

    for row in rows:
        ws.append([row.get(key) for key, _ in EXPORT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    thin = Side(style="thin", color="CBD5E1")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row_cells:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Hyperlinks for search/source columns.
    header_index = {title: idx + 1 for idx, title in enumerate(headers)}
    link_columns = ["Поиск 2GIS", "Поиск Google", "Поиск Яндекс", "Источник e-Qazyna"]
    for col_title in link_columns:
        col_idx = header_index.get(col_title)
        if not col_idx:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    # Number/date formatting.
    if "Дата создания" in header_index:
        col = header_index["Дата создания"]
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col).number_format = "yyyy-mm-dd hh:mm:ss"
    if "Время выгрузки" in header_index:
        col = header_index["Время выгрузки"]
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col).number_format = "yyyy-mm-dd hh:mm:ss"

    widths = {
        "A": 20,
        "B": 20,
        "C": 16,
        "D": 16,
        "E": 42,
        "F": 28,
        "G": 24,
        "H": 12,
        "I": 42,
        "J": 42,
        "K": 58,
        "L": 58,
        "M": 24,
        "N": 18,
        "O": 32,
        "P": 16,
        "Q": 36,
        "R": 18,
        "S": 18,
        "T": 36,
        "U": 24,
        "V": 24,
        "W": 24,
        "X": 14,
        "Y": 40,
        "Z": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if ws.max_row >= 1:
        table_ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}{max(ws.max_row, 2)}"
        table = Table(displayName="EqazynaApplications", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
